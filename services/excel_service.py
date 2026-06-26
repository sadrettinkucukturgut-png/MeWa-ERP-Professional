from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from PySide6.QtWidgets import QFileDialog, QMessageBox


class ExcelService:
    @staticmethod
    def export_excel(parent, headers: Sequence[str], rows: Sequence[Sequence[object]], filename: str, sheet_title: str, success_message: str) -> bool:
        try:
            save_path, _ = QFileDialog.getSaveFileName(
                parent,
                "Excel Dosyasını Kaydet",
                filename,
                "Excel Dosyaları (*.xlsx)",
            )
            if not save_path:
                return False

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = sheet_title[:31]

            header_font = Font(bold=True)
            sheet.append([str(value) for value in headers])
            for cell in sheet[1]:
                cell.font = header_font

            for row in rows:
                sheet.append(["" if value is None else str(value) for value in row])

            for column in range(1, len(headers) + 1):
                column_letters = get_column_letter(column)
                max_length = 10
                for cell in sheet[column_letters]:
                    if cell.value is None:
                        continue
                    max_length = max(max_length, len(str(cell.value)))
                sheet.column_dimensions[column_letters].width = min(max_length + 2, 40)

            sheet.freeze_panes = "A2"
            workbook.save(save_path)
            QMessageBox.information(parent, "Başarılı", success_message)
            return True
        except Exception as exc:  # pragma: no cover - defensive fallback
            QMessageBox.critical(parent, "Hata", f"Excel dosyası oluşturulurken bir hata oluştu:\n{exc}")
            return False
